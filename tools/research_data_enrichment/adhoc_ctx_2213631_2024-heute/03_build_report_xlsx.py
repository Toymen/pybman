# Step 3 of 3. Requires openpyxl (pip install openpyxl).
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SCR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCR, "..", "..", "..", "outputs", "research_data_enrichment")
with open(f"{OUT_DIR}/forschungsdaten_ctx_2213631_2024-heute.json") as f:
    data = json.load(f)

meta = data["meta"]
summary = data["zusammenfassung"]
pubs = data["publikationen"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

# ---------- Sheet 1: Übersicht ----------
ws1 = wb.active
ws1.title = "Übersicht"
headers1 = [
    "PuRe Item ID", "Titel", "DOI", "Datum veröffentlicht", "Genre", "Status",
    "Forschungsgruppen (Tags)", "Forschungsdaten verlinkt?", "Anzahl Links", "PuRe-Link",
]
ws1.append(headers1)
for c in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

for p in pubs:
    ws1.append([
        p["itemId"],
        p["title"],
        p["doi"] or "",
        p["datePublished"] or "",
        p["genre"],
        p["publicState"],
        ", ".join(p["forschungsgruppen_tags"]),
        "Ja" if p["hatVerlinkteForschungsdaten"] else "Nein",
        len(p["forschungsdaten"]),
        p["pureUrl"],
    ])

widths1 = [16, 55, 22, 16, 16, 12, 22, 14, 10, 42]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
tab1 = Table(displayName="Uebersicht", ref=f"A1:{get_column_letter(len(headers1))}{len(pubs)+1}")
tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws1.add_table(tab1)

# ---------- Sheet 2: Forschungsdaten (one row per link) ----------
ws2 = wb.create_sheet("Forschungsdaten")
headers2 = [
    "PuRe Item ID", "Titel", "DOI Publikation", "Forschungsgruppen (Tags)",
    "Forschungsdaten-URL", "Kategorie", "Quelle", "Anmerkung",
    "Manuelle Prüfung empfohlen",
]
ws2.append(headers2)
for c in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

row_count = 0
for p in pubs:
    for link in p["forschungsdaten"]:
        ws2.append([
            p["itemId"],
            p["title"],
            p["doi"] or "",
            ", ".join(p["forschungsgruppen_tags"]),
            link["url"],
            link["kategorie"],
            link.get("quelle", ""),
            link.get("anmerkung", ""),
            "Ja" if link.get("manuelle_pruefung_empfohlen") else "Nein",
        ])
        row_count += 1

widths2 = [16, 48, 20, 20, 46, 34, 26, 60, 14]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=2, max_row=row_count + 1):
    row[7].alignment = WRAP  # Anmerkung column wraps
ws2.freeze_panes = "A2"
tab2 = Table(displayName="Forschungsdaten", ref=f"A1:{get_column_letter(len(headers2))}{row_count+1}")
tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws2.add_table(tab2)

# ---------- Sheet 3: Forschungsgruppen (pivot-style: per tag, counts) ----------
ws3 = wb.create_sheet("Forschungsgruppen")
group_counts = {}
group_with_data = {}
for p in pubs:
    for tag in p["forschungsgruppen_tags"]:
        group_counts[tag] = group_counts.get(tag, 0) + 1
        if p["hatVerlinkteForschungsdaten"]:
            group_with_data[tag] = group_with_data.get(tag, 0) + 1

headers3 = ["Forschungsgruppe (Tag)", "Publikationen gesamt", "Publikationen mit Forschungsdaten", "Anteil"]
ws3.append(headers3)
for c in range(1, len(headers3) + 1):
    cell = ws3.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

for tag in sorted(group_counts, key=lambda t: -group_counts[t]):
    total = group_counts[tag]
    withdata = group_with_data.get(tag, 0)
    ws3.append([tag, total, withdata, f"{withdata/total:.0%}"])

widths3 = [26, 20, 30, 12]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"
tab3 = Table(displayName="Forschungsgruppen", ref=f"A1:{get_column_letter(len(headers3))}{len(group_counts)+1}")
tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws3.add_table(tab3)

# ---------- Sheet 4: Zusammenfassung & Methodik ----------
ws4 = wb.create_sheet("Zusammenfassung & Methodik", 0)
ws4.column_dimensions["A"].width = 34
ws4.column_dimensions["B"].width = 110

def add_kv(row, k, v):
    ws4.cell(row=row, column=1, value=k).font = Font(bold=True)
    cell = ws4.cell(row=row, column=2, value=v)
    cell.alignment = WRAP

r = 1
ws4.cell(row=r, column=1, value="Forschungsdaten in Pure-Publikationen").font = Font(bold=True, size=14)
r += 2
add_kv(r, "Kontext", f"{meta['kontext']} — {meta['kontextName']}"); r += 1
add_kv(r, "Zeitraum", f"{meta['zeitraum']['von']} bis {meta['zeitraum']['bis']}"); r += 1
add_kv(r, "Erzeugt am", meta["erzeugtAm"]); r += 1
add_kv(r, "Publikationen gesamt", summary["publikationenGesamt"]); r += 1
add_kv(r, "davon mit DOI", summary["publikationenMitDOI"]); r += 1
add_kv(r, "Publikationen mit verlinkten Forschungsdaten", summary["publikationenMitVerlinktenForschungsdaten"]); r += 1
add_kv(r, "davon neu durch externe Discovery gefunden (nicht in PuRe getaggt)", summary["davonNeuDurchExterneDiscoveryGefunden"]); r += 2

ws4.cell(row=r, column=1, value="Links pro Kategorie").font = Font(bold=True)
r += 1
for cat, n in summary["anzahlLinksProKategorie"].items():
    add_kv(r, cat, n); r += 1
r += 1

add_kv(r, "Quelle", meta["quelle"]); r += 1
r += 1
ws4.cell(row=r, column=1, value="Methodik").font = Font(bold=True); r += 1
add_kv(r, "1. PuRe-Tag", meta["methodik"]); r += 2
add_kv(r, "2. Externe Cross-Validierung", meta["methodik_zweiter_durchgang"]); r += 2
add_kv(r, "Einschränkung", meta["einschraenkung"]); r += 2
add_kv(r, "Berücksichtigte Kategorien", ", ".join(meta["kategorien_forschungsdaten"])); r += 1
add_kv(r, "Ausgeschlossene Kategorien", ", ".join(meta["ausgeschlossene_kategorien"])); r += 1

for rr in range(1, r + 1):
    ws4.row_dimensions[rr].height = 15

out_path = f"{OUT_DIR}/Forschungsdaten_ctx_2213631_2024-heute.xlsx"
wb.save(out_path)
print("wrote", out_path)
print("rows sheet2 (links):", row_count)
